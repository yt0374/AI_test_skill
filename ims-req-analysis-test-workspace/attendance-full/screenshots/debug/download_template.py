"""Download IMS import template and regenerate test data in correct format."""
from playwright.sync_api import sync_playwright
from pathlib import Path

TEMPLATE_DIR = Path(r"C:\Users\yanta\.claude\skills\ims-req-analysis-test-workspace\attendance-full\erp_tests\utils\extracted_data")

with sync_playwright() as p:
    browser = p.chromium.launch(headless=False)
    page = browser.new_page(viewport={"width": 1920, "height": 1080})

    # Login
    page.goto("http://test.fj.dtsimple.pro/")
    page.wait_for_load_state("networkidle")
    page.wait_for_timeout(2000)
    if "/login" in page.url:
        page.fill("#form_item_username", "admin")
        page.fill("#form_item_password", "metas2660")
        sel = page.locator(".ant-select")
        if sel.count() > 0:
            sel.first.click()
            page.wait_for_timeout(500)
            page.locator(".ant-select-item-option-content").first.click()
        page.click(".ant-btn-primary")
        page.wait_for_timeout(5000)
        page.wait_for_load_state("networkidle")

    # Go to 考勤登记 → Excel导入
    page.goto("http://test.fj.dtsimple.pro/#/personnel/workAttendance/detail")
    page.wait_for_load_state("networkidle")
    page.wait_for_timeout(3000)
    page.locator("button:has-text('excel导入')").first.click()
    page.wait_for_timeout(2000)

    # Click "下载导入模板" to get the correct template
    # Intercept download
    with page.expect_download() as download_info:
        page.locator("button:has-text('下载导入模板')").first.click()

    download = download_info.value
    template_path = TEMPLATE_DIR / f"ims_template_{download.suggested_filename}"
    download.save_as(str(template_path))
    print(f"[OK] Template downloaded: {template_path}")

    page.wait_for_timeout(1000)
    browser.close()

    # Read the template to see the format
    import openpyxl
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active
    print(f"\nTemplate headers (row 1):")
    headers = []
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col).value
        headers.append(val)
        print(f"  Col {col}: {val}")

    # Check if there's a second row (example data)
    if ws.max_row >= 2:
        print(f"\nExample row 2:")
        for col in range(1, ws.max_column + 1):
            print(f"  Col {col}: {ws.cell(row=2, column=col).value}")
